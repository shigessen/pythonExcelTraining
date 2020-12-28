import openpyxl

categorys = ((0,""),(10,"ポロシャツ"), (11,"ドレスシャツ"), (12,"カジュアルシャツ"), \
            (13,"Tシャツ"), (15,"カーディガン"),(16,"セーター"),(17,"スウェット"), \
            (18,"パーカ"))
sizes = ("コード","分類名","S","M","L","LL","XL")
#二次元リスト作成
# 以下はダメ　要素のリストがすべて同じオブジェクトになるため
#order_amount= [[0]*len(sizes)] * len(categorys)
order_amount= [[0]*len(sizes) for i in range(len(categorys))]

for j in range(len(sizes)):
    order_amount[0][j] = sizes[j]
for i in range(1,len(categorys)):
    order_amount[i][0] = categorys[i][0]
    order_amount[i][1] = categorys[i][1]
#print(order_amount)   

wb = openpyxl.load_workbook(r".\data\ordersList.xlsx")
sh = wb.active
for row in range(2, sh.max_row + 1):
    category = sh["I" + str(row)].value
    size = sh["L" + str(row)].value
    amount = sh["M" + str(row)].value
    for i in range(1,len(categorys)):
        if category == order_amount[i][0]:
            for j in range(2,len(sizes)):
                if size == order_amount[0][j]:
                    order_amount[i][j] += amount


owb = openpyxl.Workbook()
osh = owb.active
row = 1
for order_row in order_amount:
    col = 1
    size_sum = 0 
    for order_col in order_row:
        osh.cell(row, col).value = order_col
        if  row > 1 and col > 2: 
            #print(order_col)
            size_sum += order_col
        col += 1
    if row == 1:
        osh.cell(row, col).value =  "合計"
    else:
        osh.cell(row, col).value =  size_sum
    row += 1

owb.save(r".\data\orders_aggregate.xlsx")