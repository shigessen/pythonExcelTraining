import pathlib  # 標準ライブラリ
import openpyxl # 外部ライブラリ　pip install openpyxl
import csv      # 標準ライブラリ


lwb = openpyxl.Workbook()           #売上一覧表ワークブック
lsh = lwb.active                    #売上一覧表ワークシート
list_row = 1
path = pathlib.Path(r".\data\sales")    #相対パス指定
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(9,19):
                if sh.cell(dt_row, 2).value != None:
                    #より説明的なコード
                    #lsh.cell(row=list_row, column=1).value = \
                    #    sh.cell(row=2, column=7).value   #伝票NO
                    lsh.cell(list_row, 1).value = sh.cell(2, 7).value   #伝票NO
                    lsh.cell(list_row, 2).value = sh.cell(3, 7).value   #日付
                    lsh.cell(list_row, 3).value = sh.cell(4, 3).value   #得意先コード
                    lsh.cell(list_row, 4).value = sh.cell(7, 8).value   #担当者コード
                    lsh.cell(list_row, 5).value = sh.cell(dt_row, 1).value #No                    
                    lsh.cell(list_row, 6).value = sh.cell(dt_row, 2).value #商品コード 
                    lsh.cell(list_row, 7).value = sh.cell(dt_row, 3).value #商品名
                    lsh.cell(list_row, 8).value = sh.cell(dt_row, 4).value #数量
                    lsh.cell(list_row, 9).value = sh.cell(dt_row, 5).value #単価
                    lsh.cell(list_row, 10).value = sh.cell(dt_row, 4).value * \
                                                sh.cell(dt_row, 5).value #金額
                    lsh.cell(list_row, 11).value = sh.cell(dt_row, 7).value #備考                                      
                    list_row += 1

#lwb.save(r".\data\sales\salesList.xlsx")

with open(r".\data\sales\salesList.csv","w",encoding="utf_8_sig") as fp:
    writer = csv.writer(fp, lineterminator="\n")
    for row in lsh.rows:
        writer.writerow([col.value for col in row])
        print([col.value for col in row])
